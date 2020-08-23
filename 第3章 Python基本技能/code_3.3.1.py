# -*- encoding: utf-8 -*-

"""
3.3.1 读写Excel文件
"""

# 1. 使用openpyxl读写.xlsx格式的文件
# ------------------------------

# 使用openpyxl编辑Excel文件
from openpyxl import load_workbook

wb = load_workbook(r"..\res\ionosphere.xlsx")
print(wb.sheetnames) # 打印所有的表名组成的列表：['电离层']
sh = wb.get_sheet_by_name("电离层") # 选择表
print(sh.max_row) # 有效行数：351
print(sh.max_column) # 有效列数：34
print(sh['C1']) # 返回C1单元格对象
print(sh['C1'].value) # 返回C1单元格内容：0.99539
print(sh[1][2].value) # 也可以这样指定单元格：0.99539
sh['C1'].value = 99.99 # 修改单元格内容
wb.save(r"..\res\ionosphere_demo.xlsx") # 保存文件

# 使用openpyxl创建Excel文件
from  openpyxl import  Workbook

wb = Workbook() # 创建book
sh0 = wb.active # 激活默认的sheet
sha = wb.create_sheet("成绩表") # 创建新表
shb = wb.create_sheet("收支表") # 创建新表
sha.append(['姓名','语文','数学']) # 可以在末尾追加一行
sha.append(['Alice',95,99])
sha['B2'] = 98 # 也可以单独写单元格
print(wb.sheetnames) # 显示全部表名：['Sheet', '成绩表', '收支表']
del wb['Sheet'] # 删除表
wb.save(r"..\res\demo.xlsx") # 保存文件

# 设置样式
from openpyxl import  Workbook
from openpyxl.styles import Font, colors, Alignment

wb = Workbook()
sh = wb.active
f1 = Font(name='微软雅黑', size=16, italic=True, color=colors.BLACK, bold=True)
sh['A1'].font = f1 # 设置字体
align = Alignment(horizontal='center', vertical='center')
sh['B2'].alignment = align # 设置对齐方式
sh.row_dimensions[2].height = 24 # 设置第2行高度
sh.column_dimensions['C'].width = 20 # 设置C列宽度
sh.merge_cells('A3:C4') # 合并A3到C4的单元格


# 2.使用xlrd读.xls格式的文件
# ------------------------------

import xlrd

book = xlrd.open_workbook(r"..\res\ionosphere.xls")
print(book.sheet_names()) # 获取全部表名：['电离层']
sh = book.sheet_by_name('电离层') # 通过表名取得sheet对象
sh = book.sheet_by_index(0) # 通过索引取得sheet对象
print(sh.nrows) # 有效行数：351
print(sh.ncols) # 有效列数：34
print(sh.row_values(3, start_colx=3, end_colx=8)) # 读取第3行的第3列到第8列的值
print(sh.col_values(2, start_rowx=3, end_rowx=10)) # 读取第2列的第3行到第10行的值
print(sh.cell_value(3,4)) # 返回第3行第4列的值

# 3. 使用xlwt生成.xls格式的文件
# ------------------------------

import xlwt

data = [('水星',0.58,0.05), ('金星',1.08,0.82), ('地球',1.50,1.00), ('火星',2.28,0.11), ('木星',7.78,317.94), ('土星',14.27,95.18), ('天王星',28.70,14.63), ('海王星',44.97,17.22)]

book = xlwt.Workbook() # 创建book对象
sh = book.add_sheet("太阳系行星") # 添加名为"太阳系行星"的sheet
col_names = ['行星', '距离（亿千米）', '与地球的质量比']  # 列名称
for col, name in enumerate(col_names): # 列名写在第0行
    sh.write(0, col, name)
    sh.col(col).width = 256 * 20  # 设置列宽度为20个字符宽度

for i, line in enumerate(data): # 逐行逐列写入数据
    for j, item in enumerate(line):
        sh.write(i+1, j, item)

sh.write_merge(9, 9, 0, 2, xlwt.Formula('SUM(C2:C9)'))
book.save(r"..\res\planet.xls")
